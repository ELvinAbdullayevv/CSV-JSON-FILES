import csv

rows = []
with open("PKN.WA.csv", "r") as file:
    csv_reader = csv.reader(file)
    header = next(csv_reader)

    for line in csv_reader:
        rows.append(line)

print(header)
print(rows)

import openpyxl

def  excel_download(file, title):
    data = openpyxl.load_workbook(file)
    page = data[title]
    setr_sayi = page.max_column
    sutun_sayi = page.max_row

    data = []
    for i in range(1, setr_sayi):
        setir = []
        for j in range(1, sutun_sayi):
            setir.append(page.cell(i,j).value)
        data.append(setir)
    return(data)

print(excel_download("./data.xlsx", "kitaplar"))
# print(file.sheetnames)


import datetime

tday = datetime.date.today()


print(tday.weekday())

print(tday.isoweekday())


# Monday 0 - Sunday 6 (weekday)
# Monday 0 - Sunday 7 (isoweekday)


import datetime

today = datetime.date.today()

war_date = datetime.date(2025, 10, 15)

till_war = war_date - today

print(till_war)